import sys
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton, QTextEdit, QFileDialog, QMessageBox, QComboBox
import pyodbc
import openpyxl

class DatabaseImporter(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()
        self.headers = []

    def initUI(self):
        self.setWindowTitle('Access to Excel Importer')
        self.setGeometry(100, 100, 600, 500)

        main_layout = QVBoxLayout()

        # Database connection
        db_layout = QHBoxLayout()
        db_layout.addWidget(QLabel('Access Database:'))
        self.db_path = QLineEdit()
        db_layout.addWidget(self.db_path)
        self.db_browse = QPushButton('Browse')
        self.db_browse.clicked.connect(self.browse_db)
        db_layout.addWidget(self.db_browse)
        main_layout.addLayout(db_layout)

        # Table name
        table_layout = QHBoxLayout()
        table_layout.addWidget(QLabel('Table Name:'))
        self.table_name = QLineEdit()
        table_layout.addWidget(self.table_name)
        main_layout.addLayout(table_layout)

        # Record number range
        record_layout = QHBoxLayout()
        record_layout.addWidget(QLabel('NUMBER Range:'))
        self.start_number = QLineEdit()
        record_layout.addWidget(self.start_number)
        record_layout.addWidget(QLabel('to'))
        self.end_number = QLineEdit()
        record_layout.addWidget(self.end_number)
        main_layout.addLayout(record_layout)

        # Column selection for sorting
        sort_layout = QHBoxLayout()
        sort_layout.addWidget(QLabel('Sort Column 1:'))
        self.sort_column1 = QComboBox()
        sort_layout.addWidget(self.sort_column1)
        sort_layout.addWidget(QLabel('Sort Column 2:'))
        self.sort_column2 = QComboBox()
        sort_layout.addWidget(self.sort_column2)
        main_layout.addLayout(sort_layout)

        # Keywords
        keyword_layout = QHBoxLayout()
        keyword_layout.addWidget(QLabel('Keywords (comma-separated):'))
        self.keywords = QLineEdit()
        self.keywords.setText('VIDEO,AUDIO,JF,NETWORK')
        keyword_layout.addWidget(self.keywords)
        main_layout.addLayout(keyword_layout)

        # Excel file
        excel_layout = QHBoxLayout()
        excel_layout.addWidget(QLabel('Excel File:'))
        self.excel_path = QLineEdit()
        excel_layout.addWidget(self.excel_path)
        self.excel_browse = QPushButton('Browse')
        self.excel_browse.clicked.connect(self.browse_excel)
        excel_layout.addWidget(self.excel_browse)
        main_layout.addLayout(excel_layout)

        # Load Columns button
        self.load_columns_button = QPushButton('Load Columns')
        self.load_columns_button.clicked.connect(self.load_columns)
        main_layout.addWidget(self.load_columns_button)

        # Import button
        self.import_button = QPushButton('Import Data')
        self.import_button.clicked.connect(self.import_data)
        main_layout.addWidget(self.import_button)

        # Status area
        self.status = QTextEdit()
        self.status.setReadOnly(True)
        main_layout.addWidget(self.status)

        self.setLayout(main_layout)

    def browse_db(self):
        file_name, _ = QFileDialog.getOpenFileName(self, "Select Access Database", "", "Access Database (*.accdb *.mdb)")
        if file_name:
            self.db_path.setText(file_name)

    def browse_excel(self):
        file_name, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "", "Excel Files (*.xlsx)")
        if file_name:
            self.excel_path.setText(file_name)

    def load_columns(self):
        try:
            conn_str = (
                r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
                f'DBQ={self.db_path.text()};'
            )
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()

            table = self.table_name.text()
            cursor.execute(f"SELECT TOP 1 * FROM {table}")
            
            self.headers = [column[0] for column in cursor.description]
            self.sort_column1.clear()
            self.sort_column2.clear()
            self.sort_column1.addItems(self.headers)
            self.sort_column2.addItems(self.headers)

            self.status.append("Columns loaded successfully.")
        except Exception as e:
            self.status.append(f"Error loading columns: {str(e)}")

    def import_data(self):
        try:
            conn_str = (
                r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
                f'DBQ={self.db_path.text()};'
            )
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()

            start_num = int(self.start_number.text())
            end_num = int(self.end_number.text())
            table = self.table_name.text()
            cursor.execute(f"SELECT * FROM {table} WHERE NUMBER BETWEEN ? AND ? ORDER BY NUMBER", start_num, end_num)
            records = cursor.fetchall()

            if not records:
                raise ValueError(f"No records found in table {table} for NUMBER range {start_num} to {end_num}")

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Imported Data"

            ws.append(self.headers)

            for record in records:
                ws.append(list(record))

            wb.save(self.excel_path.text())

            ws = wb.active
            keywords = [kw.strip() for kw in self.keywords.text().split(',')]
            ws.auto_filter.ref = ws.dimensions
            
            sort_col1 = self.sort_column1.currentText()
            sort_col2 = self.sort_column2.currentText()
            
            if sort_col1 in self.headers:
                col1_index = self.headers.index(sort_col1)
                ws.auto_filter.add_filter_column(col1_index, keywords)
            
            if sort_col2 in self.headers and sort_col2 != sort_col1:
                col2_index = self.headers.index(sort_col2)
                ws.auto_filter.add_filter_column(col2_index, keywords)

            for keyword in keywords:
                for col in range(1, ws.max_column + 1):
                    if keyword in str(ws.cell(row=1, column=col).value).upper():
                        ws.auto_filter.add_filter_column(col - 1, [keyword])

            wb.save(self.excel_path.text())

            self.status.append(f"Data imported and sorted successfully! Imported {len(records)} records.")

        except Exception as e:
            self.status.append(f"Error: {str(e)}")

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = DatabaseImporter()
    ex.show()
    sys.exit(app.exec_())
